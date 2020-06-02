from os import listdir
from os.path import isfile, join
import pandas as pd
import openpyxl
from time import clock
import random
import math

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    def read_excel_2_dataframe(self, path):
        return pd.read_excel(path + self.ext_xlsx)

    def make_Deep_PE_input_excel(self, path, data_dict, init_arr):
        pam_seq = init_arr[0]
        add_seq1_len = init_arr[1]
        spacer_len = init_arr[2]
        add_seq2_len = init_arr[3]
        pam_len = len(pam_seq)

        for chr_key, val_dict in data_dict.items():

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            row = 1
            sheet.cell(row=row, column=1, value="INDEX")
            sheet.cell(row=row, column=2, value='Target gene name')
            sheet.cell(row=row, column=3, value='Ensembl transcript ID')
            sheet.cell(row=row, column=4, value='Target sequence')
            # sheet.cell(row=row, column=5, value='gRNA binding')
            # sheet.cell(row=row, column=6, value='PAM')

            for trnscrpt_id, vals_arr in val_dict.items():
                gene_name = ""
                dscript = vals_arr[0]

                if "gene_symbol:" in dscript:
                    gene_name = dscript[dscript.index("gene_symbol:") + len("gene_symbol:"):].split(" ")[0]
                for trgt_idx in range(1, len(vals_arr)):
                    row += 1
                    sheet.cell(row=row, column=1, value=str(row - 1))
                    sheet.cell(row=row, column=2, value=gene_name)
                    sheet.cell(row=row, column=3, value=trnscrpt_id)
                    sheet.cell(row=row, column=4, value=vals_arr[trgt_idx])
                    # sheet.cell(row=row, column=5, value=vals_arr[trgt_idx][add_seq1_len: - pam_len - add_seq2_len])
                    # sheet.cell(row=row, column=6, value=vals_arr[trgt_idx][add_seq1_len + spacer_len: - add_seq2_len])

            # workbook.save(filename=path + "deep_pe_input_chr_" + str(chr_key) + "_" + str(clock()) + self.ext_xlsx)
            workbook.save(filename=path + "deep_pe_input_chr_" + str(chr_key) + self.ext_xlsx)

    def make_cas_off_finder_input(self, result_dict, init_arr):
        pam_seq = init_arr[0]
        spacr_len = init_arr[1]
        mis_mtch_num = init_arr[2]
        sub_set_num = init_arr[3]
        path = init_arr[4]
        ref_srv_path = init_arr[5]
        init_be_list = init_arr[6]

        seq_set_pool = set()
        for chrm_key, val_dict in result_dict.items():
            for trnscrpt_id, val_arr in val_dict.items():
                val_arr_len = len(val_arr)
                for idx in range(1, val_arr_len):
                    seq_set_pool.add(val_arr[idx][0][init_be_list[1]:init_be_list[1] + init_be_list[2]])

        raw_each_sub_len = len(seq_set_pool) / sub_set_num
        each_sub_len = int(raw_each_sub_len)
        # * 1000) / 1000 : raw_each_sub_len 값이 반올림되는 경우의 error 방지
        raw_last_sub_len_to_add = int(((raw_each_sub_len - each_sub_len) * sub_set_num) * 1000) / 1000
        last_sub_len_to_add = math.ceil(raw_last_sub_len_to_add)
        print("total len : " + str(len(seq_set_pool)) + ", raw_each_sub_len : " + str(raw_each_sub_len))
        print("raw_last_sub_len_to_add : " + str(raw_last_sub_len_to_add) + " , last_sub_len_to_add : " + str(
            last_sub_len_to_add))

        for fname_idx in range(sub_set_num):
            if fname_idx == sub_set_num - 1:
                each_sub_len += last_sub_len_to_add
            tmp_sub_set = set(random.sample(seq_set_pool, each_sub_len))
            seq_set_pool -= tmp_sub_set
            print("tmp_sub_set_" + str(fname_idx) + " len : " + str(len(tmp_sub_set)))
            print("rest total len : " + str(len(seq_set_pool)))
            print(" ")
            with open(path + str(fname_idx) + self.ext_txt, 'a') as f:
                f.write(ref_srv_path + "\n")
                f.write("N" * spacr_len + pam_seq + "\n")
                for data_str in tmp_sub_set:
                    f.write(data_str + "NNN " + str(mis_mtch_num) + "\n")

    def make_Deep_PE_input_tb_txt(self, path, data_dict):
        row = 0
        with open(path + self.ext_txt, 'a') as f:
            f.write("INDEX\tTarget gene name\tEnsembl transcript ID\tTarget sequence\n")
            for chr_key, val_dict in data_dict.items():
                for trnscrpt_id, vals_arr in val_dict.items():
                    gene_name = ""
                    dscript = vals_arr[0]

                    # get gene name
                    if "gene_symbol:" in dscript:
                        gene_name = dscript[dscript.index("gene_symbol:") + len("gene_symbol:"):].split(" ")[0]
                    for trgt_idx in range(1, len(vals_arr)):
                        row += 1
                        f.write(str(row) + "\t" + gene_name + "\t" + trnscrpt_id + "\t" + vals_arr[trgt_idx] + "\n")

    def make_deep_cas9_input(self, path, dict_arr, init_arr):
        pam_seq = init_arr[0]
        add_seq1_len = init_arr[1]
        spacer_len = init_arr[2]
        add_seq2_len = init_arr[3]
        pam_len = len(pam_seq)

        # del duplicates
        tmp_set = set()
        for data_dict in dict_arr:
            for chr_key, val_dict in data_dict.items():
                for trnscrpt_id, vals_arr in val_dict.items():
                    for trgt_idx in range(1, len(vals_arr)):
                        tmp_set.add(vals_arr[trgt_idx][0])

        row = 0
        with open(path + self.ext_txt, 'a') as f:
            f.write("Target number\t30 bp target sequence (4 bp + 20 bp protospacer + PAM + 3 bp)\n")
            for seq_str in tmp_set:
                row += 1
                f.write(str(row) + "\t" + seq_str + "\n")

    def make_merge_tab_txt(self, result_path, init_dict_arr, init_len):
        full_dict = init_dict_arr[0]
        abe_score_dict = init_dict_arr[1]
        cbe_score_dict = init_dict_arr[2]
        cs9_score_dict = init_dict_arr[3]

        pam_seq = init_len[0]
        add_seq1_len = init_len[1]
        spacer_len = init_len[2]
        add_seq2_len = init_len[3]
        clvg_site = init_len[4]
        pam_len = len(pam_seq)

        row = 0
        with open(result_path + self.ext_txt, 'a') as f:
            f.write("index\tchromosome\tTarget gene name\tDescription\tEnsembl transcript ID\tEnsembl Gene ID\tStrand\torder sgRNA Target sequence\torder Target context sequence\torder PAM\tcleavage site\tDeepCas9 score\tABE score\tCBE score\n")
            for chr_key, trnscrpt_val in full_dict.items():
                for trnscrpt_id, vals_arr in trnscrpt_val.items():
                    full_description = vals_arr[0]
                    for trgt_idx in range(1, len(vals_arr)):
                        cntxt_seq = vals_arr[trgt_idx][0]
                        clvg_site = vals_arr[trgt_idx][1]
                        cas9_score = 0
                        abe_score = 0
                        cbe_score = 0

                        if cntxt_seq in cs9_score_dict:
                            cas9_score = cs9_score_dict[cntxt_seq]
                        else:
                            print(cntxt_seq + " doesn't have CAS9 score")
                        if cntxt_seq in abe_score_dict:
                            abe_score = abe_score_dict[cntxt_seq]
                        else:
                            print(cntxt_seq + " doesn't have ABE score")
                        if cntxt_seq in cbe_score_dict:
                            cbe_score = cbe_score_dict[cntxt_seq]
                        else:
                            print(cntxt_seq + " doesn't have CBE score")

                        row += 1
                        # f.write(str(row) + "\t" + chr_key + "\t" + "gene_name" + "\t" + "description" + "\t" + trnscrpt_id + "\t" + "gene_id" + "\t" + "strand" + "\t" + cntxt_seq[add_seq1_len:-add_seq2_len] + "\t" + cntxt_seq + "\t" + cntxt_seq[add_seq1_len + spacer_len:-add_seq2_len] + "\t" + str(clvg_site) + "\t" + str(cas9_score) + "\t" + str(abe_score) + "\t" + str(cbe_score) + "\n")
                        f.write(str(row) + "\t" + chr_key + "\t" + full_description + "\t" + trnscrpt_id + "\t" + "gene_id" + "\t" + "strand" + "\t" + cntxt_seq[add_seq1_len:-add_seq2_len] + "\t" + cntxt_seq + "\t" + cntxt_seq[add_seq1_len + spacer_len:-add_seq2_len] + "\t" + str(clvg_site) + "\t" + str(cas9_score) + "\t" + str(abe_score) + "\t" + str(cbe_score) + "\n")

    def make_merge_excel_by_chr(self, result_path, init_dict_arr, init_len):
        full_dict = init_dict_arr[0]
        abe_score_dict = init_dict_arr[1]
        cbe_score_dict = init_dict_arr[2]
        cs9_score_dict = init_dict_arr[3]

        pam_seq = init_len[0]
        add_seq1_len = init_len[1]
        spacer_len = init_len[2]
        add_seq2_len = init_len[3]
        clvg_site = init_len[4]
        pam_len = len(pam_seq)

        for chr_key, trnscrpt_val in full_dict.items():
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            row = 1
            sheet.cell(row=row, column=1, value="index")
            sheet.cell(row=row, column=2, value='Target gene name')
            sheet.cell(row=row, column=3, value='Description')
            sheet.cell(row=row, column=4, value='Ensembl transcript ID')
            sheet.cell(row=row, column=5, value='Ensembl Gene ID')
            sheet.cell(row=row, column=6, value='Strand')
            sheet.cell(row=row, column=7, value='order sgRNA Target sequence')
            sheet.cell(row=row, column=8, value='order Target context sequence')
            sheet.cell(row=row, column=9, value='order PAM')
            sheet.cell(row=row, column=10, value='cleavage site')
            sheet.cell(row=row, column=11, value='DeepCas9 score')
            sheet.cell(row=row, column=12, value='ABE score')
            sheet.cell(row=row, column=13, value='CBE score')
            for trnscrpt_id, vals_arr in trnscrpt_val.items():
                full_description = vals_arr[0]
                full_description_arr = full_description.split(" ")
                gene_id = full_description_arr[3].replace("gene:", "")
                strand = "+"
                if ":-1" in full_description_arr[2]:
                    strand = "-"

                gene_nm = ""
                description = ""
                if "gene_symbol:" in full_description:
                    if "description:" in full_description:
                        gene_nm = full_description[
                                  full_description.index("gene_symbol:") + len("gene_symbol:"):full_description.index(
                                      "description:")]
                        description = full_description[full_description.index("description:") + len("description:"):]
                    else:
                        gene_nm = full_description[
                                  full_description.index("gene_symbol:") + len("gene_symbol:"):]

                for trgt_idx in range(1, len(vals_arr)):
                        cntxt_seq = vals_arr[trgt_idx][0]
                        clvg_site = vals_arr[trgt_idx][1]
                        cas9_score = 0
                        abe_score = 0
                        cbe_score = 0

                        if cntxt_seq in cs9_score_dict:
                            cas9_score = cs9_score_dict[cntxt_seq]
                        else:
                            print(cntxt_seq + " doesn't have CAS9 score")
                        if cntxt_seq in abe_score_dict:
                            abe_score = abe_score_dict[cntxt_seq]
                        else:
                            print(cntxt_seq + " doesn't have ABE score")
                        if cntxt_seq in cbe_score_dict:
                            cbe_score = cbe_score_dict[cntxt_seq]
                        else:
                            print(cntxt_seq + " doesn't have CBE score")
                        row += 1
                        sheet.cell(row=row, column=1, value=str(row - 1))
                        sheet.cell(row=row, column=2, value=gene_nm)
                        sheet.cell(row=row, column=3, value=description)
                        sheet.cell(row=row, column=4, value=trnscrpt_id)
                        sheet.cell(row=row, column=5, value=gene_id)
                        sheet.cell(row=row, column=6, value=strand)
                        sheet.cell(row=row, column=7, value=cntxt_seq[add_seq1_len:-add_seq2_len])
                        sheet.cell(row=row, column=8, value=cntxt_seq)
                        sheet.cell(row=row, column=9, value=cntxt_seq[add_seq1_len + spacer_len:-add_seq2_len])
                        sheet.cell(row=row, column=10, value=clvg_site)
                        sheet.cell(row=row, column=11, value=cas9_score)
                        sheet.cell(row=row, column=12, value=abe_score)
                        sheet.cell(row=row, column=13, value=cbe_score)
            workbook.save(filename=result_path + "_" + str(chr_key) + self.ext_xlsx)